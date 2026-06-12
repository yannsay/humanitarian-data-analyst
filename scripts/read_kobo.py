#!/usr/bin/env python3
"""
read_kobo.py — parse a Kobo/ODK XLSForm and write a normalised JSON cache.

Sheet names and formats vary across deployments. The script therefore works in
two phases:

  Phase 1 (--list-sheets):  Read and print the workbook's sheet names as JSON.
                             Pass the output to the LLM so it can identify which
                             sheet is the survey, which is choices, and (optionally)
                             which is a data/response sheet.

  Phase 2 (default):        Read the sheets the LLM identified and write the cache.
                             Pass --survey-sheet and --choices-sheet to override the
                             defaults ('survey' / 'choices').

  Phase 3 (query modes):    Query an existing kobo_<slug>.json cache without re-parsing
                             the xlsx. Use these in Step 3 to pull targeted slices instead
                             of reading the whole cache into context.

Usage:
    # Phase 1 — inspect sheets
    python3 scripts/read_kobo.py <dataset>.xlsx --list-sheets

    # Phase 2 — parse and write cache (structural rows excluded by default)
    python3 scripts/read_kobo.py <dataset>.xlsx --slug <slug> \\
        [--survey-sheet "Survey"] [--choices-sheet "Choices"] \\
        [--include-structural]   # include begin_group, note, calculate, etc.

    # Phase 3 query modes (read from cache, not xlsx):
    python3 scripts/read_kobo.py --cache kobo_<slug>.json --summary
    python3 scripts/read_kobo.py --cache kobo_<slug>.json --names q85,q87,q88
    python3 scripts/read_kobo.py --cache kobo_<slug>.json --group wash
    python3 scripts/read_kobo.py --cache kobo_<slug>.json --names q85 --with-choices

CONTEXT DISCIPLINE (Step 3):
    The agent MUST use --summary first (cheap orientation), then --names or --group
    for the specific variables each indicator needs. It MUST NOT read the whole
    kobo_<slug>.json into context. The --summary output is ~200 bytes; --names for
    2-3 variables is ~1-2 KB. Reading the whole file is ~100 KB — 50× too large.

Phase 2 output: kobo_<slug>.json beside the xlsx (or -o path).

JSON shape (Phase 2):
  {
    "source_file": "<filename>",
    "sheet_names": [...],
    "survey_sheet": "Survey",
    "choices_sheet": "Choices",
    "n_total": 133,                # null if no data sheet
    "skipped_rows": 0,
    "survey": [                    # substantive rows only (structural excluded by default)
      {"type": "select_one foo", "name": "var_name", "label": "Question text",
       "relevant": "...", "list_name": "foo"}
      # empty fields omitted
    ],
    "choices": {
      "list_name": [{"name": "opt", "label": "Option label"}, ...]
    }
  }
"""
import argparse
import json
import os
import re
import sys


def _require_openpyxl():
    """Import openpyxl lazily — only needed for Phase 1/2 (xlsx parsing)."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        sys.exit("read_kobo.py requires openpyxl: pip3 install openpyxl --break-system-packages")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_sheet(wb, name):
    """Return sheet by exact name (case-insensitive). Raises KeyError if absent."""
    lower_map = {s.lower(): s for s in wb.sheetnames}
    key = name.lower()
    if key not in lower_map:
        available = ", ".join(f"'{s}'" for s in wb.sheetnames)
        sys.exit(f"read_kobo.py: sheet '{name}' not found. Available: {available}")
    return wb[lower_map[key]]


def _header_map(ws):
    """Return {col_index: header_text} from the first non-empty row."""
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(c is not None for c in row):
            return {i: (str(c).strip() if c is not None else "") for i, c in enumerate(row)}
    return {}


def _find_col(headers, *candidates):
    """First header index matching any candidate (case-insensitive)."""
    lower = {v.lower(): k for k, v in headers.items()}
    for c in candidates:
        if c.lower() in lower:
            return lower[c.lower()]
    return None


def _label_col(headers):
    """Detect the English label column: 'label', 'label::English*', or first 'label*'."""
    for idx, h in headers.items():
        if h.lower() == "label":
            return idx
    for idx, h in headers.items():
        if re.match(r"label::english", h.lower()):
            return idx
    for idx, h in headers.items():
        if h.lower().startswith("label"):
            return idx
    return None


# ---------------------------------------------------------------------------
# Phase 1 — list sheets
# ---------------------------------------------------------------------------

def list_sheets(xlsx_path):
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result = {
        "source_file": os.path.basename(xlsx_path),
        "sheet_names": wb.sheetnames,
        "instruction": (
            "Pass --survey-sheet and --choices-sheet to read_kobo.py, "
            "optionally --data-sheet for n_total. "
            "Omit --data-sheet if no response dataset is present."
        )
    }
    print(json.dumps(result, indent=2))


# ---------------------------------------------------------------------------
# Phase 2 — parse
# ---------------------------------------------------------------------------

def parse_kobo(xlsx_path, slug, survey_sheet_name, choices_sheet_name, data_sheet_name,
               include_structural=False):
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    # ---- survey sheet ----
    survey_ws = _get_sheet(wb, survey_sheet_name)
    rows = list(survey_ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("read_kobo.py: survey sheet is empty")

    headers = _header_map(survey_ws)
    type_col  = _find_col(headers, "type")
    name_col  = _find_col(headers, "name")
    label_col = _label_col(headers)
    rel_col   = _find_col(headers, "relevant")

    if type_col is None or name_col is None:
        sys.exit(f"read_kobo.py: survey sheet '{survey_sheet_name}' missing 'type' or 'name' column. "
                 f"Columns found: {list(headers.values())}")

    STRUCTURAL = {"begin_group", "end_group", "begin_repeat", "end_repeat",
                  "note", "calculate", "hidden", "start", "end", "deviceid",
                  "simserial", "phonenumber", "username", "audit"}

    survey_rows = []
    skipped = 0
    for row in rows[1:]:
        try:
            type_val = str(row[type_col]).strip() if row[type_col] is not None else ""
            name_val = str(row[name_col]).strip() if row[name_col] is not None else ""
            if not type_val and not name_val:
                continue  # blank row

            label_val = ""
            if label_col is not None and label_col < len(row) and row[label_col] is not None:
                label_val = str(row[label_col]).strip()
            relevant_val = ""
            if rel_col is not None and rel_col < len(row) and row[rel_col] is not None:
                relevant_val = str(row[rel_col]).strip()

            list_name = None
            m = re.match(r"select_(one|multiple)\s+(\S+)", type_val, re.IGNORECASE)
            if m:
                list_name = m.group(2)

            base_type = type_val.split()[0].lower() if type_val else ""
            is_structural = base_type in STRUCTURAL

            # Skip structural rows unless --include-structural requested (D12 slim)
            if is_structural and not include_structural:
                continue

            # Build a slim entry — omit empty fields (D12 context fix)
            entry: dict = {"type": type_val, "name": name_val}
            if label_val:
                entry["label"] = label_val
            if relevant_val:
                entry["relevant"] = relevant_val
            if list_name:
                entry["list_name"] = list_name
            # Keep is_structural only when including structural rows (for caller awareness)
            if include_structural:
                entry["is_structural"] = is_structural
            survey_rows.append(entry)
        except Exception:
            skipped += 1

    # ---- choices sheet ----
    choices = {}
    if choices_sheet_name:
        choices_ws = _get_sheet(wb, choices_sheet_name)
        ch_rows = list(choices_ws.iter_rows(values_only=True))
        if ch_rows:
            ch_headers = _header_map(choices_ws)
            ln_col = _find_col(ch_headers, "list_name", "list name")
            cn_col = _find_col(ch_headers, "name")
            cl_col = _label_col(ch_headers)

            if ln_col is not None and cn_col is not None:
                for row in ch_rows[1:]:
                    try:
                        ln  = str(row[ln_col]).strip() if row[ln_col]  is not None else ""
                        cn  = str(row[cn_col]).strip() if row[cn_col]  is not None else ""
                        clv = ""
                        if cl_col is not None and cl_col < len(row) and row[cl_col] is not None:
                            clv = str(row[cl_col]).strip()
                        if ln and cn:
                            choices.setdefault(ln, []).append({"name": cn, "label": clv})
                    except Exception:
                        pass

    # ---- data/response sheet — count rows (optional) ----
    n_total = None
    if data_sheet_name:
        data_ws = _get_sheet(wb, data_sheet_name)
        data_rows = list(data_ws.iter_rows(values_only=True))
        if len(data_rows) > 1:
            n_total = sum(1 for r in data_rows[1:] if any(c is not None for c in r))
        else:
            n_total = 0

    result = {
        "source_file": os.path.basename(xlsx_path),
        "sheet_names": sheet_names,
        "survey_sheet": survey_sheet_name,
        "choices_sheet": choices_sheet_name,
        "n_total": n_total,
        "skipped_rows": skipped,
        "survey": survey_rows,
        "choices": choices,
    }
    if slug:
        result["slug"] = slug
    return result


# ---------------------------------------------------------------------------
# Phase 3 — query modes (read from existing cache JSON)
# ---------------------------------------------------------------------------

def load_cache(cache_path: str) -> dict:
    with open(cache_path, encoding="utf-8") as f:
        return json.load(f)


def query_summary(cache: dict) -> dict:
    """Return a cheap orientation summary: module groups + question counts.
    This is the ONLY query that should run before knowing which variables to look up.
    """
    survey = cache.get("survey", [])
    # Attempt to infer groups from name prefixes (e.g. q85 → group by leading alpha)
    # Use a simple heuristic: questions with same first word in name or label prefix
    groups: dict[str, list[str]] = {}
    current_group = "_root"
    for row in survey:
        name = row.get("name", "")
        # Detect group boundaries via original structural info if available
        if row.get("_group"):
            current_group = row["_group"]
        groups.setdefault(current_group, []).append(name)

    substantive = [r for r in survey if not r.get("is_structural", False)]
    return {
        "source_file": cache.get("source_file"),
        "n_total": cache.get("n_total"),
        "survey_question_count": len(substantive),
        "choice_list_count": len(cache.get("choices", {})),
        "all_question_names": [r["name"] for r in substantive],
        "note": "Use --names q1,q2 or --group <name> to retrieve specific questions with their choice lists.",
    }


def query_names(cache: dict, names: list[str], with_choices: bool = True) -> dict:
    """Return specific survey entries (+ their choice lists if with_choices)."""
    name_set = set(names)
    survey = cache.get("survey", [])
    matched = [r for r in survey if r.get("name") in name_set]

    result: dict = {"questions": matched}
    if with_choices:
        choices_needed = set()
        for r in matched:
            if r.get("list_name"):
                choices_needed.add(r["list_name"])
        choices = cache.get("choices", {})
        result["choices"] = {k: choices[k] for k in choices_needed if k in choices}
    return result


def query_group(cache: dict, group_name: str, with_choices: bool = True) -> dict:
    """Return all survey entries whose _group field matches group_name (case-insensitive)."""
    survey = cache.get("survey", [])
    g = group_name.lower()
    matched = [r for r in survey if r.get("_group", "").lower() == g
               or r.get("name", "").lower().startswith(g)]

    result: dict = {"group": group_name, "questions": matched}
    if with_choices:
        choices_needed = {r["list_name"] for r in matched if r.get("list_name")}
        choices = cache.get("choices", {})
        result["choices"] = {k: choices[k] for k in choices_needed if k in choices}
    return result


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description=(
            "Phase 1: --list-sheets. "
            "Phase 2: parse xlsx to cache. "
            "Phase 3: query cache with --cache + (--summary | --names | --group)."
        )
    )
    # Phase 2 positional (optional so Phase 3 works without it)
    ap.add_argument("xlsx", nargs="?", default=None,
                    help="Path to the Kobo/ODK .xlsx file (Phase 1 & 2 only)")
    ap.add_argument("--list-sheets", action="store_true",
                    help="Phase 1: print sheet names as JSON and exit")
    ap.add_argument("--slug", default=None,
                    help="Short identifier (used in output filename)")
    ap.add_argument("--survey-sheet", default="survey",
                    help="Name of the survey sheet (default: 'survey')")
    ap.add_argument("--choices-sheet", default="choices",
                    help="Name of the choices sheet (default: 'choices')")
    ap.add_argument("--data-sheet", default=None,
                    help="Name of the response/data sheet for n_total (omit if absent)")
    ap.add_argument("--include-structural", action="store_true",
                    help="Include structural rows (begin_group, note, calculate, …) in cache")
    ap.add_argument("-o", "--out", default=None,
                    help="Output JSON path (default: kobo_<slug>.json beside the xlsx)")
    # Phase 3 query flags
    ap.add_argument("--cache", default=None,
                    help="Phase 3: path to an existing kobo_<slug>.json cache to query")
    ap.add_argument("--summary", action="store_true",
                    help="Phase 3: return cheap orientation summary (question names + counts)")
    ap.add_argument("--names", default=None,
                    help="Phase 3: comma-separated variable names to retrieve, e.g. q85,q87")
    ap.add_argument("--group", default=None,
                    help="Phase 3: return all questions in this survey group/module")
    ap.add_argument("--no-choices", action="store_true",
                    help="Phase 3: omit choice lists from --names / --group output")
    a = ap.parse_args()

    # ---- Phase 3: query mode ----
    if a.cache:
        cache = load_cache(a.cache)
        with_choices = not a.no_choices
        if a.summary:
            result = query_summary(cache)
        elif a.names:
            names = [n.strip() for n in a.names.split(",") if n.strip()]
            result = query_names(cache, names, with_choices=with_choices)
        elif a.group:
            result = query_group(cache, a.group, with_choices=with_choices)
        else:
            ap.error("--cache requires one of --summary, --names, or --group")
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    # ---- Phase 1 ----
    if a.xlsx is None:
        ap.error("xlsx path is required unless using --cache")
    if a.list_sheets:
        list_sheets(a.xlsx)
        return

    # ---- Phase 2 ----
    slug = a.slug or os.path.splitext(os.path.basename(a.xlsx))[0]
    out_path = a.out or os.path.join(
        os.path.dirname(os.path.abspath(a.xlsx)), f"kobo_{slug}.json"
    )

    data = parse_kobo(
        a.xlsx, slug,
        survey_sheet_name=a.survey_sheet,
        choices_sheet_name=a.choices_sheet,
        data_sheet_name=a.data_sheet,
        include_structural=a.include_structural,
    )

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    n_survey = len(data["survey"])
    print(f"wrote {out_path}")
    print(f"  n_total={data['n_total']}  survey_questions={n_survey}  "
          f"choice_lists={len(data['choices'])}  skipped_rows={data['skipped_rows']}")


if __name__ == "__main__":
    main()
